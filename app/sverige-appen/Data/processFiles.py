import pandas as pd
from openpyxl import load_workbook
import os


####
#This is a script for converting an .xlsx from the national emission database to a format that we can use for our sqlite database.
#TODO:
# - add support for iterating over all files
#    - Currently, some files have more descriptive rows than others. Maybe locate the word 'Huvudsektor', and make the removing of rows relative to that. 

####

# READ File

def open_files():
    path = "./Nationella-emmisions-databasen"

    frames = []
    for filename in os.listdir(path):
        print("reading" + filename)
        file_path = os.path.join(path, filename)
        with open(file_path, "rb") as file:
            sheet_name = load_workbook(file).sheetnames[0].strip()
            data = pd.read_excel(file, sheet_name=sheet_name)
            frames.append(process_file(data, sheet_name))
         

    with open("./data.csv", "w", encoding="utf-8") as file:
        df = pd.concat(frames)

        df.to_csv(file, lineterminator='\n', index=False, encoding='utf-8')
 

def process_file(data, sheet_name):
    destination_row = data[data.iloc[:, 0] == "Huvudsektor"].index[0]
    print(destination_row)
    year_row = destination_row - 1
    #They are at E to O
    year_cols = [i for i in range (4, 15)]

    data.iloc[destination_row, year_cols] = data.iloc[year_row, year_cols].values
    data.columns = data.iloc[destination_row]
    data.drop(data.index[0:destination_row + 1], inplace=True)
    data.reset_index(drop=True, inplace=True)
    #Ignore sectors
    all_sectors = data[(data['Huvudsektor'] == 'Alla') & (data['Undersektor'] == 'Alla')]
    all_sectors.drop(columns=['Huvudsektor', 'Undersektor'], inplace=True)
    melted = all_sectors.melt(id_vars=['Län', 'Kommun'], var_name='År', value_name='Value')
    melted['Emission type'] = sheet_name
    return melted

if __name__ == "__main__":
    open_files()
  